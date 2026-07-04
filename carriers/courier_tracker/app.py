"""
Multi-Courier Tracking Software v2.0 — Anti-Block Edition
Dark theme GUI | 200-300 simultaneous | No hanging | Smart fallback
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
import json, os, csv, time, re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from tracker_engine import TrackingEngine, detect_courier, TrackResult

CONFIG_FILE = "config.json"
DEFAULT_CONFIG = {
    "fedex_api_key":"","fedex_secret_key":"",
    "dhl_api_key":"",
    "ups_api_key":"","ups_client_id":"","ups_client_secret":"",
    "aramex_username":"","aramex_password":"","aramex_account_no":"",
    "aramex_account_pin":"","aramex_account_entity":"","aramex_country_code":"IN",
    "shiprocket_email":"","shiprocket_password":"",
    "onpoint_api_key":"","shipglobal_api_key":"",
    "xindus_api_key":"","xindus_client_id":"",
}

# ── COLORS ──────────────────────────────────────────────────────────
BG=    "#0d1117"; CARD=  "#161b22"; CARD2= "#21262d"
BORD=  "#30363d"; BLUE=  "#58a6ff"; GREEN= "#3fb950"
ORANGE="#e3b341"; RED=   "#f85149"; PURPLE="#bc8cff"
CYAN=  "#39d353"; TEXT=  "#c9d1d9"; MUTED= "#8b949e"
WHITE= "#ffffff";  YELLOW="#ffa657"

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                return {**DEFAULT_CONFIG, **json.load(f)}
        except Exception:
            pass
    return DEFAULT_CONFIG.copy()

def save_config(cfg):
    with open(CONFIG_FILE,"w") as f:
        json.dump(cfg, f, indent=2)

# ── SCREEN-ADAPTIVE SIZING ───────────────────────────────────────────
def fit_to_screen(win, min_w=1050, min_h=680, max_w=1800, max_h=1040,
                   width_frac=0.85, height_frac=0.85):
    """Size + center the window as a percentage of the ACTUAL screen
    resolution, so it fits properly on small laptops as well as large
    monitors instead of always opening at a fixed 1450x860."""
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    w = max(min_w, min(int(sw * width_frac), max_w, sw - 40))
    h = max(min_h, min(int(sh * height_frac), max_h, sh - 60))
    x, y = (sw - w) // 2, (sh - h) // 2
    win.geometry(f"{w}x{h}+{x}+{y}")
    return w, h

# ── STYLES ──────────────────────────────────────────────────────────
def setup_styles():
    s = ttk.Style(); s.theme_use("clam")
    s.configure("TFrame",        background=BG)
    s.configure("Card.TFrame",   background=CARD)
    s.configure("TLabel",        background=BG,    foreground=TEXT,   font=("Segoe UI",10))
    s.configure("Card.TLabel",   background=CARD,  foreground=TEXT,   font=("Segoe UI",10))
    s.configure("H.TLabel",      background=CARD,  foreground=BLUE,   font=("Segoe UI",12,"bold"))
    s.configure("TButton",       background=BLUE,  foreground=WHITE,  font=("Segoe UI",10,"bold"), borderwidth=0, focuscolor="none")
    s.map("TButton",             background=[("active","#388bfd"),("pressed","#1f6feb")])
    s.configure("G.TButton",     background=GREEN,  foreground=WHITE,  font=("Segoe UI",10,"bold"))
    s.map("G.TButton",           background=[("active","#2ea043")])
    s.configure("R.TButton",     background=RED,    foreground=WHITE,  font=("Segoe UI",10,"bold"))
    s.map("R.TButton",           background=[("active","#da3633")])
    s.configure("O.TButton",     background=ORANGE, foreground=WHITE,  font=("Segoe UI",10,"bold"))
    s.map("O.TButton",           background=[("active","#d29922")])
    s.configure("TEntry",        fieldbackground=CARD2, foreground=TEXT, insertcolor=TEXT, font=("Segoe UI",10))
    s.configure("TCombobox",     fieldbackground=CARD2, foreground=TEXT, selectbackground=BLUE, font=("Segoe UI",10))
    s.map("TCombobox",           fieldbackground=[("readonly",CARD2)])
    s.configure("TNotebook",     background=BG, borderwidth=0)
    s.configure("TNotebook.Tab", background=CARD, foreground=MUTED, font=("Segoe UI",10), padding=[16,8])
    s.map("TNotebook.Tab",       background=[("selected",CARD2)], foreground=[("selected",BLUE)])
    s.configure("Treeview",         background=CARD, foreground=TEXT, fieldbackground=CARD, rowheight=26, font=("Segoe UI",9))
    s.configure("Treeview.Heading", background=CARD2, foreground=BLUE, font=("Segoe UI",9,"bold"), relief="flat")
    s.map("Treeview",               background=[("selected",BLUE)], foreground=[("selected",WHITE)])
    s.configure("TProgressbar",  troughcolor=CARD2, background=GREEN, thickness=5)
    s.configure("TScrollbar",    background=CARD2, troughcolor=CARD, arrowcolor=MUTED, borderwidth=0)

# ── MAIN APP ────────────────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root     = root
        self.config   = load_config()
        self.results  = []
        self.stop_evt = threading.Event()
        self.running  = False
        self._cfg_entries = {}
        setup_styles()
        self._win()
        self._ui()

    def _win(self):
        self.root.title("🚀 Multi-Courier Tracker v2.0 | Anti-Block Edition")
        fit_to_screen(self.root)   # auto-size + center per actual screen resolution
        self.root.minsize(1050,680)
        self.root.resizable(True, True)
        self.root.configure(bg=BG)
        self.root.protocol("WM_DELETE_WINDOW", lambda: (self.stop_evt.set(), self.root.destroy()))

    # ── UI SKELETON ─────────────────────────────────────────────────
    def _ui(self):
        # HEADER
        hf = tk.Frame(self.root, bg=CARD, height=54)
        hf.pack(fill="x"); hf.pack_propagate(False)
        tk.Label(hf, text="🚀 MULTI-COURIER TRACKER", bg=CARD, fg=BLUE,
                 font=("Segoe UI",15,"bold")).pack(side="left", padx=18, pady=10)
        tk.Label(hf, text="FedEx · DHL · UPS · Aramex · Shiprocket · OnPoint · ShipGlobal · Xindus",
                 bg=CARD, fg=MUTED, font=("Segoe UI",9)).pack(side="left", pady=14)
        self._clock_lbl = tk.Label(hf, text="", bg=CARD, fg=MUTED, font=("Segoe UI",9))
        self._clock_lbl.pack(side="right", padx=18)
        self._tick()

        # NOTEBOOK
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True)
        t1,t2,t3,t4 = (ttk.Frame(self.nb) for _ in range(4))
        self.nb.add(t1, text="  📦 Track  ")
        self.nb.add(t2, text="  📊 Results  ")
        self.nb.add(t3, text="  ⚙️ API Keys  ")
        self.nb.add(t4, text="  ❓ Help  ")
        self._tab_track(t1)
        self._tab_results(t2)
        self._tab_config(t3)
        self._tab_help(t4)

        # STATUS BAR
        sb = tk.Frame(self.root, bg=CARD2, height=26)
        sb.pack(fill="x", side="bottom"); sb.pack_propagate(False)
        self._status = tk.Label(sb, text="Ready — Tracking numbers paste karein aur Track karein",
                                 bg=CARD2, fg=MUTED, font=("Segoe UI",9))
        self._status.pack(side="left", padx=12, pady=3)
        self._count = tk.Label(sb, text="0 results", bg=CARD2, fg=BLUE, font=("Segoe UI",9,"bold"))
        self._count.pack(side="right", padx=12, pady=3)
        self._method_lbl = tk.Label(sb, text="", bg=CARD2, fg=GREEN, font=("Segoe UI",8))
        self._method_lbl.pack(side="right", padx=20, pady=3)

    def _tick(self):
        self._clock_lbl.configure(text=datetime.now().strftime("%d %b %Y  %H:%M:%S"))
        self.root.after(1000, self._tick)

    # ── TAB: TRACK ──────────────────────────────────────────────────
    def _tab_track(self, f):
        # LEFT panel
        lp = tk.Frame(f, bg=CARD, width=400)
        lp.pack(side="left", fill="y", padx=(14,6), pady=14)
        lp.pack_propagate(False)

        tk.Label(lp, text="TRACKING NUMBERS", bg=CARD, fg=BLUE,
                 font=("Segoe UI",10,"bold")).pack(anchor="w", padx=12, pady=(14,2))
        tk.Label(lp, text="Ek line = ek number  |  Format: COURIER:AWB ya sirf AWB",
                 bg=CARD, fg=MUTED, font=("Segoe UI",8)).pack(anchor="w", padx=12, pady=(0,6))

        self._inp = scrolledtext.ScrolledText(lp, bg=CARD2, fg=TEXT, insertbackground=TEXT,
            font=("Consolas",9), relief="flat", wrap="word", height=20,
            selectbackground=BLUE, highlightthickness=1, highlightbackground=BORD)
        self._inp.pack(fill="both", expand=True, padx=12, pady=(0,6))
        self._inp.insert("end",
            "# Yahan tracking numbers likhein:\n"
            "# Auto detect:\n# 799999999999\n# 1Z999AA10123456784\n\n"
            "# Courier specify karein:\n# FedEx:799999999999\n# DHL:1234567890\n# UPS:1Z999AA10123456784\n# Aramex:12345678901\n")

        # Controls
        cf = tk.Frame(lp, bg=CARD); cf.pack(fill="x", padx=12, pady=(0,6))
        tk.Label(cf, text="Default:", bg=CARD, fg=MUTED, font=("Segoe UI",9)).pack(side="left")
        self._courier = tk.StringVar(value="Auto Detect")
        cb = ttk.Combobox(cf, textvariable=self._courier, width=14, state="readonly",
            values=["Auto Detect","FedEx","DHL","UPS","Aramex","Shiprocket","OnPoint","ShipGlobal","Xindus"])
        cb.pack(side="right")

        # Buttons
        bf = tk.Frame(lp, bg=CARD); bf.pack(fill="x", padx=12, pady=(0,4))
        self._btn_go = ttk.Button(bf, text="🚀  START TRACKING", command=self._start)
        self._btn_go.pack(fill="x", pady=(0,5), ipady=3)
        self._btn_stop = ttk.Button(bf, text="⛔  STOP", style="R.TButton",
                                     command=self._stop, state="disabled")
        self._btn_stop.pack(fill="x", pady=(0,5))

        r2 = tk.Frame(bf, bg=CARD); r2.pack(fill="x")
        ttk.Button(r2, text="📂 Import", style="O.TButton",
                   command=self._import).pack(side="left", fill="x", expand=True, padx=(0,3))
        ttk.Button(r2, text="🗑 Clear",
                   command=lambda: self._inp.delete("1.0","end")).pack(side="right", fill="x", expand=True, padx=(3,0))

        # Progress
        pf = tk.Frame(lp, bg=CARD); pf.pack(fill="x", padx=12, pady=(8,4))
        self._prog_lbl = tk.Label(pf, text="0 / 0", bg=CARD, fg=MUTED, font=("Segoe UI",9))
        self._prog_lbl.pack(anchor="w", pady=(0,3))
        self._progbar = ttk.Progressbar(pf, mode="determinate")
        self._progbar.pack(fill="x")

        # RIGHT panel — live log + KPIs
        rp = tk.Frame(f, bg=BG); rp.pack(fill="both", expand=True, padx=(0,14), pady=14)
        tk.Label(rp, text="LIVE LOG", bg=BG, fg=BLUE,
                 font=("Segoe UI",10,"bold")).pack(anchor="w", pady=(0,6))

        # KPI row
        kf = tk.Frame(rp, bg=BG); kf.pack(fill="x", pady=(0,8))
        self._kpis = {}
        for lbl, key, col in [("Total","total",BLUE),("✅ Done","done",GREEN),
                                ("📦 Delivered","del",CYAN),("❌ Error","err",RED),
                                ("🔀 Method","method",ORANGE)]:
            c = tk.Frame(kf, bg=CARD, padx=14, pady=8)
            c.pack(side="left", fill="x", expand=True, padx=(0,6))
            tk.Label(c, text=lbl, bg=CARD, fg=MUTED, font=("Segoe UI",8)).pack()
            v = tk.StringVar(value="0" if key!="method" else "-")
            self._kpis[key] = v
            tk.Label(c, textvariable=v, bg=CARD, fg=col,
                     font=("Segoe UI",20,"bold")).pack()

        self._log = scrolledtext.ScrolledText(rp, bg=CARD, fg=TEXT, font=("Consolas",9),
            relief="flat", state="disabled", wrap="word",
            highlightthickness=1, highlightbackground=BORD)
        self._log.pack(fill="both", expand=True)
        for tag, col in [("g",GREEN),("r",RED),("b",BLUE),("o",ORANGE),("m",MUTED),("c",CYAN)]:
            self._log.tag_config(tag, foreground=col)

    # ── TAB: RESULTS ────────────────────────────────────────────────
    def _tab_results(self, f):
        tb = tk.Frame(f, bg=CARD, height=48); tb.pack(fill="x", padx=14, pady=(10,0)); tb.pack_propagate(False)
        tk.Label(tb, text="Filter:", bg=CARD, fg=MUTED, font=("Segoe UI",9)).pack(side="left", padx=(12,4), pady=10)
        self._flt = tk.StringVar(value="All")
        for txt,val in [("All","All"),("✅ Delivered","Delivered"),("🚚 Transit","Transit"),
                         ("❌ Error","Error"),("🔑 API Missing","API"),("❓ Unknown","Unknown")]:
            tk.Radiobutton(tb, text=txt, variable=self._flt, value=val,
                bg=CARD, fg=TEXT, selectcolor=CARD2, activebackground=CARD,
                font=("Segoe UI",9), command=self._filter).pack(side="left", padx=5, pady=10)
        self._srch = tk.StringVar()
        self._srch.trace_add("write", lambda *a: self._filter())
        ttk.Entry(tb, textvariable=self._srch, width=18).pack(side="left", padx=(12,0), pady=10)
        ttk.Button(tb, text="💾 Excel", style="G.TButton", command=self._export_excel).pack(side="right", padx=(4,12), pady=8)
        ttk.Button(tb, text="📄 CSV", style="O.TButton", command=self._export_csv).pack(side="right", padx=4, pady=8)

        cols = ("#","AWB","Courier","Status","Location","Description","Last Update","Method","Del?")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", selectmode="browse")
        ws = {"#":38,"AWB":165,"Courier":90,"Status":140,"Location":145,"Description":215,"Last Update":130,"Method":70,"Del?":60}
        for c in cols:
            self.tree.heading(c, text=c, command=lambda _c=c: self._sort(_c))
            self.tree.column(c, width=ws[c], minwidth=35,
                anchor="center" if c in ("#","Del?","Method") else "w")
        vsb = ttk.Scrollbar(f, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(f, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(14,0), pady=(6,0))
        vsb.pack(side="right", fill="y", pady=(6,0), padx=(0,4))
        hsb.pack(side="bottom", fill="x", padx=14, pady=(0,6))
        self.tree.tag_configure("del",  background="#0d2818", foreground=GREEN)
        self.tree.tag_configure("err",  background="#2d1117", foreground=RED)
        self.tree.tag_configure("api",  background="#1c1a0d", foreground=YELLOW)
        self.tree.tag_configure("trn",  background="#0d1f2d", foreground=CYAN)
        self.tree.bind("<Double-1>", self._detail)

    # ── TAB: CONFIG ─────────────────────────────────────────────────
    def _tab_config(self, f):
        canvas = tk.Canvas(f, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(f, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG)
        win = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))

        tk.Label(inner, text="API Keys & Credentials", bg=BG, fg=BLUE,
                 font=("Segoe UI",13,"bold")).pack(anchor="w", padx=20, pady=(18,3))
        tk.Label(inner, text="Courier accounts se API keys lekar yahan fill karein. Bina API key ke bhi software kaam karta hai (public fallback).",
                 bg=BG, fg=MUTED, font=("Segoe UI",9)).pack(anchor="w", padx=20, pady=(0,14))

        sections = [
            ("🟣 FedEx",          [("fedex_api_key","API Key"),("fedex_secret_key","Secret Key")]),
            ("🟡 DHL",            [("dhl_api_key","API Key")]),
            ("🟤 UPS",            [("ups_client_id","Client ID"),("ups_client_secret","Client Secret"),("ups_api_key","Access Key")]),
            ("🔴 Aramex",         [("aramex_username","Username"),("aramex_password","Password*"),
                                    ("aramex_account_no","Account No"),("aramex_account_pin","Account PIN*"),
                                    ("aramex_account_entity","Account Entity"),("aramex_country_code","Country Code")]),
            ("🔴 Shiprocket",     [("shiprocket_email","Login Email"),("shiprocket_password","Password*")]),
            ("🔵 OnPoint/17Track",[("onpoint_api_key","API Key")]),
            ("🟢 ShipGlobal",     [("shipglobal_api_key","API Key")]),
            ("🔵 Xindus",         [("xindus_api_key","API Key"),("xindus_client_id","Client ID")]),
        ]
        for title, fields in sections:
            sec = tk.Frame(inner, bg=CARD, pady=2); sec.pack(fill="x", padx=20, pady=(0,10))
            tk.Label(sec, text=title, bg=CARD, fg=TEXT, font=("Segoe UI",11,"bold")).pack(anchor="w", padx=14, pady=(10,5))
            for key, label in fields:
                rf = tk.Frame(sec, bg=CARD); rf.pack(fill="x", padx=14, pady=2)
                tk.Label(rf, text=label+":", bg=CARD, fg=MUTED, font=("Segoe UI",9), width=18, anchor="w").pack(side="left")
                show = "*" if "*" in label else ""
                e = ttk.Entry(rf, show=show, width=48, font=("Segoe UI",9))
                e.insert(0, self.config.get(key,""))
                e.pack(side="left", padx=(0,8))
                self._cfg_entries[key] = e
            tk.Frame(sec, bg=BORD, height=1).pack(fill="x", padx=14, pady=(5,0))

        bfr = tk.Frame(inner, bg=BG); bfr.pack(fill="x", padx=20, pady=14)
        ttk.Button(bfr, text="💾  Save API Keys", style="G.TButton",
                   command=self._save_cfg).pack(side="left", ipadx=14, ipady=4)
        tk.Label(bfr, text="  Settings config.json mein save hote hain",
                 bg=BG, fg=MUTED, font=("Segoe UI",8)).pack(side="left", padx=10)
        tk.Label(inner, text="⚡ Tip: Bina API key ke bhi software public tracking se data lene ki koshish karta hai.",
                 bg=BG, fg=ORANGE, font=("Segoe UI",9)).pack(anchor="w", padx=20, pady=(0,16))

    # ── TAB: HELP ───────────────────────────────────────────────────
    def _tab_help(self, f):
        t = scrolledtext.ScrolledText(f, bg=CARD, fg=TEXT, font=("Segoe UI",10),
                                      relief="flat", wrap="word", padx=22, pady=18)
        t.pack(fill="both", expand=True, padx=14, pady=14)
        t.insert("1.0", HELP_TEXT)
        t.configure(state="disabled")

    # ── TRACKING LOGIC ──────────────────────────────────────────────
    def _parse_input(self):
        text = self._inp.get("1.0","end").strip()
        default = self._courier.get()
        if default == "Auto Detect": default = "Auto"
        items = []
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("#"): continue
            if ":" in line:
                parts = line.split(":",1)
                cou, tn = parts[0].strip(), parts[1].strip()
            else:
                tn, cou = line, default
            if tn: items.append({"tracking_no": tn, "courier": cou})
        return items

    def _start(self):
        if self.running: return
        items = self._parse_input()
        if not items:
            messagebox.showwarning("Koi Number Nahi","Pehle tracking numbers likhein!")
            return
        self.running = True
        self.stop_evt.clear()
        self.results = []
        self._clear_tree()
        self._btn_go.configure(state="disabled")
        self._btn_stop.configure(state="normal")
        self._progbar["value"] = 0
        self._progbar["maximum"] = len(items)
        for k in self._kpis: self._kpis[k].set("0" if k!="method" else "-")
        self._kpis["total"].set(str(len(items)))
        self._log_msg(f"🚀 Tracking shuru — {len(items)} shipments (Anti-block mode ON)", "b")

        stats = {"done":0,"del":0,"err":0,"api":0,"scrape":0,"public":0}

        def progress_cb(done, total, r: TrackResult):
            if self.stop_evt.is_set(): return
            stats["done"] = done
            if r.delivered:   stats["del"]   += 1
            if r.error:       stats["err"]   += 1
            m = r.method_used
            if m == "API":    stats["api"]   += 1
            elif m == "Scrape": stats["scrape"] += 1
            elif m == "Public": stats["public"] += 1

            tag = "g" if r.delivered else "r" if r.error else "c"
            self._log_msg(
                f"[{done}/{total}] {r.tracking_no} ({r.courier}) → {(r.status or '')[:38]}  [{r.method_used}]", tag)

            def _ui():
                self._progbar["value"] = done
                self._prog_lbl.configure(text=f"{done} / {total}")
                self._kpis["done"].set(str(done))
                self._kpis["del"].set(str(stats["del"]))
                self._kpis["err"].set(str(stats["err"]))
                self._kpis["method"].set(f"API:{stats['api']} W:{stats['scrape']} P:{stats['public']}")
                self._add_row(r)
                self._count.configure(text=f"{done} results")
            self.root.after(0, _ui)

        def run():
            engine = TrackingEngine(self.config)
            res = engine.track_bulk(items, progress_callback=progress_cb, stop_event=self.stop_evt)
            self.results = res
            def _done():
                self.running = False
                self._btn_go.configure(state="normal")
                self._btn_stop.configure(state="disabled")
                self._progbar["value"] = len(items)
                self._log_msg(
                    f"✅ Complete! {len(items)} tracked — {stats['del']} delivered, {stats['err']} errors | "
                    f"API:{stats['api']} Web:{stats['scrape']} Public:{stats['public']}", "g")
                self._status.configure(
                    text=f"✅ Done — {len(items)} tracked | {stats['del']} delivered | {stats['err']} errors")
                self._method_lbl.configure(
                    text=f"API:{stats['api']} | Web:{stats['scrape']} | Public:{stats['public']}")
                self.nb.select(1)
            self.root.after(0, _done)

        threading.Thread(target=run, daemon=True).start()

    def _stop(self):
        self.stop_evt.set()
        self.running = False
        self._btn_go.configure(state="normal")
        self._btn_stop.configure(state="disabled")
        self._log_msg("⛔ Tracking stopped by user.", "o")

    def _log_msg(self, msg, tag=""):
        def _do():
            self._log.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            self._log.insert("end", f"[{ts}] {msg}\n", tag)
            self._log.see("end")
            self._log.configure(state="disabled")
        self.root.after(0, _do)

    def _status_tag(self, r: TrackResult):
        if r.delivered: return "del"
        if r.error and "API" in (r.error or ""): return "api"
        if r.error: return "err"
        return "trn"

    def _add_row(self, r: TrackResult):
        d = "✅" if r.delivered else "-"
        s = (r.status or r.error or "Unknown")[:38]
        self.tree.insert("","end",
            values=(len(self.tree.get_children()), r.tracking_no, r.courier,
                    s, r.location[:38], r.description[:50],
                    r.timestamp, r.method_used, d),
            tags=(self._status_tag(r),))

    def _clear_tree(self):
        for i in self.tree.get_children(): self.tree.delete(i)

    def _filter(self):
        flt  = self._flt.get()
        srch = self._srch.get().lower()
        self._clear_tree()
        for r in self.results:
            if not r: continue
            ok = True
            if flt=="Delivered"  and not r.delivered: ok=False
            elif flt=="Transit"  and (r.delivered or r.error): ok=False
            elif flt=="Error"    and not r.error: ok=False
            elif flt=="API"      and "API" not in (r.error or ""): ok=False
            elif flt=="Unknown"  and r.status not in ("Unknown Courier","Check Website"): ok=False
            if srch and not any(srch in s.lower() for s in
                [r.tracking_no, r.courier, r.status or "", r.location or ""]): ok=False
            if ok: self._add_row(r)

    def _sort(self, col):
        items = [(self.tree.set(c, col), c) for c in self.tree.get_children("")]
        items.sort()
        for i,(_, c) in enumerate(items): self.tree.move(c,"",i)

    def _detail(self, _evt):
        sel = self.tree.selection()
        if not sel: return
        tn  = self.tree.item(sel[0],"values")[1]
        r   = next((x for x in self.results if x and x.tracking_no==tn), None)
        if not r: return
        w = tk.Toplevel(self.root)
        w.title(f"Detail — {tn}")
        fit_to_screen(w, min_w=560, min_h=420, max_w=900, max_h=700,
                      width_frac=0.45, height_frac=0.55)
        w.resizable(True, True)
        w.configure(bg=BG)
        tk.Label(w, text=f"  {r.tracking_no}", bg=BG, fg=BLUE,
                 font=("Segoe UI",14,"bold")).pack(anchor="w", padx=18, pady=(16,2))
        tk.Label(w, text=f"  Courier: {r.courier}  |  Status: {r.status}  |  Method: {r.method_used}",
                 bg=BG, fg=TEXT, font=("Segoe UI",10)).pack(anchor="w", padx=18, pady=(0,8))
        txt = scrolledtext.ScrolledText(w, bg=CARD, fg=TEXT, font=("Consolas",9),
                                         relief="flat", wrap="word")
        txt.pack(fill="both", expand=True, padx=18, pady=(0,18))
        lines = [
            f"Tracking No  : {r.tracking_no}",
            f"Courier      : {r.courier}",
            f"Status       : {r.status}",
            f"Location     : {r.location}",
            f"Description  : {r.description}",
            f"Last Update  : {r.timestamp}",
            f"ETA          : {r.eta}",
            f"Delivered    : {'✅ Yes' if r.delivered else 'No'}",
            f"Method Used  : {r.method_used}",
            "", "─── TRACKING EVENTS ───────────────────────────────────────",
        ] + (r.events or ["No events — API key add karein ya website par check karein"])
        if r.error: lines += ["", f"⚠️  Error: {r.error}"]
        txt.insert("1.0", "\n".join(lines))
        txt.configure(state="disabled")

    def _import(self):
        path = filedialog.askopenfilename(title="Import",
            filetypes=[("CSV/Excel","*.csv *.xlsx *.xls"),("All","*.*")])
        if not path: return
        lines = []
        try:
            if path.endswith(".csv"):
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        tn  = (row.get("tracking_no") or row.get("Tracking No") or row.get("AWB","")).strip()
                        cou = (row.get("courier") or row.get("Courier","")).strip()
                        if tn: lines.append(f"{cou}:{tn}" if cou else tn)
            else:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                hdr  = [str(c or "").lower() for c in rows[0]]
                ti   = next((i for i,h in enumerate(hdr) if "track" in h or "awb" in h), 0)
                ci   = next((i for i,h in enumerate(hdr) if "courier" in h or "carrier" in h), -1)
                for row in rows[1:]:
                    tn  = str(row[ti] or "").strip()
                    cou = str(row[ci] or "").strip() if ci>=0 else ""
                    if tn and tn!="None": lines.append(f"{cou}:{tn}" if cou else tn)
                wb.close()
            self._inp.delete("1.0","end")
            self._inp.insert("end", "\n".join(lines))
            self._log_msg(f"📂 Imported {len(lines)} tracking numbers", "b")
        except Exception as e:
            messagebox.showerror("Import Error", str(e))

    def _save_cfg(self):
        for k, e in self._cfg_entries.items():
            self.config[k] = e.get().strip()
        save_config(self.config)
        messagebox.showinfo("✅ Saved","API keys saved!\nAb tracking karein.")

    def _export_excel(self):
        if not self.results:
            messagebox.showwarning("⚠️","Pehle tracking karein!"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"Tracking_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        if not path: return
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title="Tracking Results"
            hdrs = ["#","Tracking No","Courier","Status","Location","Description",
                    "Last Update","ETA","Method","Delivered","Error"]
            hfill = PatternFill("solid", fgColor="0D1117")
            hfont = Font(bold=True, color="58A6FF", name="Segoe UI", size=10)
            thin  = Side(style="thin", color="30363D")
            brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill=hfill; c.font=hfont
                c.alignment=Alignment(horizontal="center",vertical="center")
            fills = {
                "del":  PatternFill("solid", fgColor="0D2818"),
                "err":  PatternFill("solid", fgColor="2D1117"),
                "api":  PatternFill("solid", fgColor="1C1A0D"),
                "trn":  PatternFill("solid", fgColor="0D1F2D"),
            }
            for ri, r in enumerate(self.results, 2):
                if not r: continue
                tag = self._status_tag(r)
                fl  = fills.get(tag, PatternFill("solid", fgColor="161B22"))
                vals = [ri-1, r.tracking_no, r.courier, r.status, r.location,
                        r.description, r.timestamp, r.eta, r.method_used,
                        "Yes" if r.delivered else "No", r.error or ""]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=v)
                    cell.fill=fl; cell.border=brd
                    cell.font=Font(name="Segoe UI", size=9,
                                   color="3FB950" if r.delivered else "F85149" if r.error else "C9D1D9")
                    cell.alignment=Alignment(vertical="center")
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, w in enumerate([5,22,13,22,20,35,18,14,10,10,25], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            messagebox.showinfo("✅ Exported", f"Saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_csv(self):
        if not self.results:
            messagebox.showwarning("⚠️","Pehle tracking karein!"); return
        path = filedialog.asksaveasfilename(defaultextension=".csv",
            filetypes=[("CSV","*.csv")],
            initialfile=f"Tracking_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")
        if not path: return
        try:
            with open(path,"w",newline="",encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=list(TrackResult("","").to_dict().keys()))
                w.writeheader()
                for r in self.results:
                    if r: w.writerow(r.to_dict())
            messagebox.showinfo("✅ Exported", f"Saved:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))


# ── HELP TEXT ───────────────────────────────────────────────────────
HELP_TEXT = """
🚀 MULTI-COURIER TRACKER v2.0 — ANTI-BLOCK EDITION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚡ ANTI-BLOCK FEATURES (Naya):
─────────────────────────────
✅ 3-Layer Fallback System:
   Layer 1: Official API (agar key hai)
   Layer 2: Web Scraping (API nahi to)
   Layer 3: Public endpoints (last resort)

✅ Rate Limiting: Har courier ke liye alag speed control
✅ Rotating User-Agents: Bot detection se bachav
✅ Smart Retry: 3 attempts + exponential backoff
✅ Session Pooling: Connection reuse for speed
✅ Per-courier semaphores: Same courier pe 6 se zyada parallel calls nahi

📦 TRACKING NUMBERS FORMAT:
────────────────────────────
Auto detect (recommended):
  799999999999
  1Z999AA10123456784
  1234567890

Courier specify karein:
  FedEx:799999999999
  DHL:1234567890
  UPS:1Z999AA10123456784
  Aramex:12345678901
  Shiprocket:SR123456789
  OnPoint:OP123456789
  ShipGlobal:SG123456789
  Xindus:XI123456789

📊 STATUS COLUMN (Method):
────────────────────────────
API    → Official API se data mila (fastest, most accurate)
Scrape → Website se data scrape kiya (medium)
Public → Public endpoint use kiya (basic data)
Failed → Koi bhi method kaam nahi kiya

⚙️ API KEYS (Optional but recommended):
────────────────────────────────────────
FedEx:      developer.fedex.com
DHL:        developer.dhl.com
UPS:        developer.ups.com
Aramex:     aramex.com (business account)
Shiprocket: app.shiprocket.in (login email+pass)
OnPoint:    17track.net/en/api (free plan available)
ShipGlobal: shipglobal.in
Xindus:     xindus.co

💡 TIPS:
─────────
• API key nahi hai? Koi baat nahi — public fallback use hoga
• 200-300 numbers: 25 threads parallel, ~30-60 seconds mein done
• Rate limit error aane par software automatically wait karta hai
• Method column dekhein — API > Scrape > Public (accuracy order)
• Double-click any row → full tracking history

🔧 TROUBLESHOOTING:
────────────────────
"Check Website" → Courier ne block kiya, manually verify karein
"HTTP 429"      → Too many requests, software auto-retry karega
"HTTP 401"      → API key galat/expire, update karein
"Timeout"       → Internet slow hai ya courier server down
"""

# ── LAUNCH ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
