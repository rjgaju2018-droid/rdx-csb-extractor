"""
RDX Solution — CSB-V Extractor (Complete App)
===============================================
Flash Screen  +  Login Page  +  Dashboard (stats)  +  Processing Window
(CPU/RAM analog dial monitor + Add Shipping Bills + Excel export)

UI Libraries:
    pip install customtkinter pillow psutil

Extractor Libraries:
    pip install pdfplumber pymupdf pytesseract openpyxl pandas

Run:
    python rdx_csb_app.py
"""

# ══════════════════════════════════════════════════════════════════════════════
#  STANDARD IMPORTS
# ══════════════════════════════════════════════════════════════════════════════
import os, re, sys, glob, json, math, hashlib, threading, time
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from PIL import Image, ImageTk

try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
try:
    _RESAMPLE = Image.Resampling.LANCZOS   # Pillow 10+
except AttributeError:
    _RESAMPLE = Image.LANCZOS              # Pillow <10

# ── Extractor imports ────────────────────────────────────────────────────────
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import fitz
    import pytesseract
    if os.name == "nt":
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# ── Auto-updater (GitHub Releases based) ────────────────────────────────────
from version import __version__
import updater

# ══════════════════════════════════════════════════════════════════════════════
#  THEME & COLOURS  (RDX Gold Palette)
# ══════════════════════════════════════════════════════════════════════════════
CREAM       = "#F0EBE0"
DARK_BG     = "#1A1510"
CARD_BG     = "#211C14"
GOLD_DARK   = "#8B6914"
GOLD_MID    = "#C9962A"
GOLD_LIGHT  = "#E8C44A"
GOLD_SHINE  = "#F5D97A"
WHITE       = "#FFFFFF"
GREY_TEXT   = "#A09070"
RED_ERR     = "#E05555"
GREEN_OK    = "#55B055"

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

# ══════════════════════════════════════════════════════════════════════════════
#  USER STORE  (JSON file based)
# ══════════════════════════════════════════════════════════════════════════════
USER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rdx_users.json")

def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def load_users() -> dict:
    if os.path.exists(USER_FILE):
        try:
            with open(USER_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    default = {"admin": {"password": _hash("admin123"), "email": "admin@rdxsolution.com"}}
    save_users(default)
    return default

def save_users(users: dict):
    with open(USER_FILE, "w") as f:
        json.dump(users, f, indent=2)

# ══════════════════════════════════════════════════════════════════════════════
#  STATS STORE  (Dashboard Data)
# ══════════════════════════════════════════════════════════════════════════════
STATS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rdx_stats.json")

def load_stats() -> dict:
    default = {"total_runs": 0, "total_files_processed": 0,
               "total_rows_extracted": 0, "last_run": None,
               "last_run_files": 0, "last_run_rows": 0, "last_output": None}
    if os.path.exists(STATS_FILE):
        try:
            with open(STATS_FILE) as f:
                data = json.load(f)
                default.update(data)
        except Exception:
            pass
    return default

def save_stats(stats: dict):
    with open(STATS_FILE, "w") as f:
        json.dump(stats, f, indent=2)

def record_run(files_count: int, rows_count: int, output_path: str):
    stats = load_stats()
    stats["total_runs"] += 1
    stats["total_files_processed"] += files_count
    stats["total_rows_extracted"] += rows_count
    stats["last_run"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
    stats["last_run_files"] = files_count
    stats["last_run_rows"] = rows_count
    stats["last_output"] = output_path
    save_stats(stats)
    return stats

# ══════════════════════════════════════════════════════════════════════════════
#  LOGO HELPER
# ══════════════════════════════════════════════════════════════════════════════
LOGO_PATHS = [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "rdx_soloution_logo.png"),
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.png"),
]

def load_logo(size=(220, 90)) -> ImageTk.PhotoImage | None:
    for p in LOGO_PATHS:
        if os.path.exists(p):
            try:
                img = Image.open(p).convert("RGBA").resize(size, _RESAMPLE)
                return ImageTk.PhotoImage(img)
            except Exception:
                pass
    return None

def load_ctk_logo(size=(220, 90)) -> ctk.CTkImage | None:
    for p in LOGO_PATHS:
        if os.path.exists(p):
            try:
                img = Image.open(p).convert("RGBA")
                return ctk.CTkImage(img, size=size)
            except Exception:
                pass
    return None

# ══════════════════════════════════════════════════════════════════════════════
#  ROUND DIAL GAUGE  (analog watch style — CPU / RAM monitor)
# ══════════════════════════════════════════════════════════════════════════════
class DialGauge(tk.Canvas):
    def __init__(self, parent, label="CPU", size=140, **kwargs):
        super().__init__(parent, width=size, height=size,
                         bg=CARD_BG, highlightthickness=0, **kwargs)
        self.size = size
        self.label = label
        self.value = 0
        self._draw(0)

    def _draw(self, value):
        self.delete("all")
        s = self.size
        cx, cy = s / 2, s / 2
        r = s * 0.42
        start_ang, sweep_ang = 135, 270

        self.create_oval(cx - r - 8, cy - r - 8, cx + r + 8, cy + r + 8,
                         outline=GOLD_DARK, width=2, fill=DARK_BG)
        self.create_arc(cx - r, cy - r, cx + r, cy + r,
                         start=start_ang, extent=-sweep_ang,
                         style="arc", outline="#3A3322", width=8)
        
        if value < 60:
            col = GREEN_OK
        elif value < 85:
            col = GOLD_MID
        else:
            col = RED_ERR
            
        val_extent = -sweep_ang * (value / 100.0)
        if value > 0:
            self.create_arc(cx - r, cy - r, cx + r, cy + r,
                             start=start_ang, extent=val_extent,
                             style="arc", outline=col, width=8)
                             
        for i in range(11):
            ang = math.radians(start_ang - sweep_ang * (i / 10))
            x1 = cx + (r - 10) * math.cos(ang)
            y1 = cy - (r - 10) * math.sin(ang)
            x2 = cx + (r - 2) * math.cos(ang)
            y2 = cy - (r - 2) * math.sin(ang)
            self.create_line(x1, y1, x2, y2, fill=GREY_TEXT, width=1)
            
        needle_ang = math.radians(start_ang - sweep_ang * (value / 100.0))
        nx = cx + (r - 14) * math.cos(needle_ang)
        ny = cy - (r - 14) * math.sin(needle_ang)
        self.create_line(cx, cy, nx, ny, fill=col, width=3, capstyle="round")
        self.create_oval(cx - 5, cy - 5, cx + 5, cy + 5, fill=GOLD_LIGHT, outline="")

        self.create_text(cx, cy + r * 0.45, text=f"{value:.0f}%",
                         fill=WHITE, font=("Segoe UI", 13, "bold"))
        self.create_text(cx, cy + r * 0.45 + 18, text=self.label,
                         fill=GREY_TEXT, font=("Segoe UI", 9))

    def set_value(self, value):
        self.value = max(0, min(100, value))
        self._draw(self.value)

# ══════════════════════════════════════════════════════════════════════════════
#  FLASH / SPLASH SCREEN
# ══════════════════════════════════════════════════════════════════════════════
class SplashScreen(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.overrideredirect(True)
        self.configure(bg=DARK_BG)

        W, H = 540, 320
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

        border = tk.Frame(self, bg=GOLD_MID, bd=0)
        border.place(relx=0, rely=0, relwidth=1, relheight=1)
        inner = tk.Frame(border, bg=DARK_BG)
        inner.place(relx=0.003, rely=0.006, relwidth=0.994, relheight=0.988)

        logo_img = load_logo(size=(260, 105))
        if logo_img:
            self._logo_ref = logo_img
            lbl = tk.Label(inner, image=logo_img, bg=DARK_BG)
            lbl.place(relx=0.5, rely=0.32, anchor="center")
        else:
            tk.Label(inner, text="RDx Solution", font=("Georgia", 28, "bold"),
                     fg=GOLD_LIGHT, bg=DARK_BG).place(relx=0.5, rely=0.32, anchor="center")

        tk.Label(inner, text="Courier Shipping Bill  ·  Data Extractor",
                 font=("Segoe UI", 11), fg=GREY_TEXT, bg=DARK_BG
                 ).place(relx=0.5, rely=0.62, anchor="center")

        self.bar_bg = tk.Frame(inner, bg=GOLD_DARK, height=4, width=380)
        self.bar_bg.place(relx=0.5, rely=0.80, anchor="center")
        self.bar_fill = tk.Frame(inner, bg=GOLD_LIGHT, height=4, width=0)
        self.bar_fill.place(x=(W - 380) // 2, rely=0.80, anchor="w")

        self.status = tk.Label(inner, text="Initializing…",
                               font=("Segoe UI", 9), fg=GREY_TEXT, bg=DARK_BG)
        self.status.place(relx=0.5, rely=0.89, anchor="center")

        self._steps = [
            (0.25, "Loading configuration…"),
            (0.55, "Preparing modules…"),
            (0.80, "Almost ready…"),
            (1.00, "Welcome to RDx Solution!"),
        ]
        self._step_idx = 0
        self.after(300, self._advance)
        self.after(4000, self._failsafe_close)

    def _failsafe_close(self):
        try:
            if self.winfo_exists():
                self.destroy()
        except Exception:
            pass

    def _advance(self):
        if self._step_idx >= len(self._steps):
            self.destroy()
            return
        frac, msg = self._steps[self._step_idx]
        self.bar_fill.configure(width=int(380 * frac))
        self.status.configure(text=msg)
        self._step_idx += 1
        self.after(600, self._advance)

# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN PAGE
# ══════════════════════════════════════════════════════════════════════════════
class LoginPage(ctk.CTkFrame):
    def __init__(self, master, on_login_success):
        super().__init__(master, fg_color=DARK_BG)
        self.on_login_success = on_login_success
        self.users = load_users()
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        left = ctk.CTkFrame(self, fg_color=CARD_BG, corner_radius=0, width=260)
        left.grid(row=0, column=0, sticky="nsew")
        left.grid_propagate(False)
        left.grid_rowconfigure(0, weight=1)
        left.grid_columnconfigure(0, weight=1)

        logo_ctk = load_ctk_logo(size=(200, 82))
        if logo_ctk:
            ctk.CTkLabel(left, image=logo_ctk, text="").grid(row=0, column=0, pady=(0, 10))
        else:
            ctk.CTkLabel(left, text="RDx\nSolution", font=("Georgia", 26, "bold"), text_color=GOLD_LIGHT).grid(row=0, column=0)

        ctk.CTkLabel(left, text="Shipping Bill\nData Extractor", font=("Segoe UI", 12), text_color=GREY_TEXT).grid(row=1, column=0, pady=(0, 40))

        self.dt_label = ctk.CTkLabel(left, text="", font=("Consolas", 10), text_color=GOLD_DARK)
        self.dt_label.grid(row=2, column=0, pady=(0, 20))
        self._tick()

        right = ctk.CTkFrame(self, fg_color=DARK_BG, corner_radius=0)
        right.grid(row=0, column=1, sticky="nsew", padx=50)
        right.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(right, text="Welcome Back", font=("Georgia", 22, "bold"), text_color=GOLD_LIGHT).grid(row=0, column=0, pady=(60, 4), sticky="w")
        ctk.CTkLabel(right, text="Sign in to continue", font=("Segoe UI", 11), text_color=GREY_TEXT).grid(row=1, column=0, sticky="w", pady=(0, 30))

        ctk.CTkLabel(right, text="Username", font=("Segoe UI", 11), text_color=GREY_TEXT).grid(row=2, column=0, sticky="w")
        self.user_entry = ctk.CTkEntry(right, placeholder_text="Enter username", height=42, fg_color=CARD_BG, border_color=GOLD_DARK, border_width=1, text_color=WHITE, font=("Segoe UI", 12))
        self.user_entry.grid(row=3, column=0, sticky="ew", pady=(4, 14))

        ctk.CTkLabel(right, text="Password", font=("Segoe UI", 11), text_color=GREY_TEXT).grid(row=4, column=0, sticky="w")
        self.pass_entry = ctk.CTkEntry(right, placeholder_text="Enter password", show="●", height=42, fg_color=CARD_BG, border_color=GOLD_DARK, border_width=1, text_color=WHITE, font=("Segoe UI", 12))
        self.pass_entry.grid(row=5, column=0, sticky="ew", pady=(4, 6))
        self.pass_entry.bind("<Return>", lambda e: self._login())

        self.show_pw = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(right, text="Show password", variable=self.show_pw, command=self._toggle_pw, font=("Segoe UI", 10), text_color=GREY_TEXT, fg_color=GOLD_DARK, hover_color=GOLD_MID, checkmark_color=WHITE, border_color=GOLD_DARK).grid(row=6, column=0, sticky="w", pady=(0, 4))

        fp_btn = ctk.CTkButton(right, text="Forgot Password?", fg_color="transparent", text_color=GOLD_MID, hover_color=CARD_BG, font=("Segoe UI", 10, "underline"), height=20, command=self._forgot_pw)
        fp_btn.grid(row=6, column=0, sticky="e")

        self.err_label = ctk.CTkLabel(right, text="", text_color=RED_ERR, font=("Segoe UI", 10))
        self.err_label.grid(row=7, column=0, pady=(4, 0))

        self.login_btn = ctk.CTkButton(right, text="LOGIN", height=44, corner_radius=6, fg_color=GOLD_MID, hover_color=GOLD_DARK, text_color=DARK_BG, font=("Segoe UI", 13, "bold"), command=self._login)
        self.login_btn.grid(row=8, column=0, sticky="ew", pady=(16, 10))

        div = ctk.CTkFrame(right, fg_color=GOLD_DARK, height=1)
        div.grid(row=9, column=0, sticky="ew", pady=10)

        ctk.CTkLabel(right, text="New user?", font=("Segoe UI", 10), text_color=GREY_TEXT).grid(row=10, column=0, pady=(0, 2))
        ctk.CTkButton(right, text="Create New Account", height=40, corner_radius=6, fg_color="transparent", border_color=GOLD_DARK, border_width=1, text_color=GOLD_LIGHT, hover_color=CARD_BG, font=("Segoe UI", 11), command=self._new_account).grid(row=11, column=0, sticky="ew", pady=(0, 40))

    def _tick(self):
        now = datetime.now()
        day, date, time_str = now.strftime("%A"), now.strftime("%d %b %Y"), now.strftime("%I:%M:%S %p")
        self.dt_label.configure(text=f"{day}\n{date}\n{time_str}")
        self.after(1000, self._tick)

    def _toggle_pw(self):
        self.pass_entry.configure(show="" if self.show_pw.get() else "●")

    def _login(self):
        uname, pw = self.user_entry.get().strip(), self.pass_entry.get()
        self.users = load_users()
        if not uname or not pw:
            self.err_label.configure(text="⚠ Username and password are required.")
            return
        if uname not in self.users:
            self.err_label.configure(text="✗ Username not found.")
            return
        if self.users[uname]["password"] != _hash(pw):
            self.err_label.configure(text="✗ Incorrect password.")
            return
        self.err_label.configure(text="")
        self.on_login_success(uname)

    def _forgot_pw(self):
        win = ctk.CTkToplevel(self)
        win.title("Reset Password")
        win.geometry("380x300")
        win.configure(fg_color=DARK_BG)
        win.grab_set()

        ctk.CTkLabel(win, text="Reset Password", font=("Georgia", 16, "bold"), text_color=GOLD_LIGHT).pack(pady=(24, 4))
        ctk.CTkLabel(win, text="Enter username and new password", font=("Segoe UI", 10), text_color=GREY_TEXT).pack(pady=(0, 16))

        u = ctk.CTkEntry(win, placeholder_text="Username", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        u.pack(padx=30, fill="x", pady=4)
        p1 = ctk.CTkEntry(win, placeholder_text="New Password", show="●", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        p1.pack(padx=30, fill="x", pady=4)
        p2 = ctk.CTkEntry(win, placeholder_text="Confirm New Password", show="●", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        p2.pack(padx=30, fill="x", pady=4)
        err = ctk.CTkLabel(win, text="", text_color=RED_ERR, font=("Segoe UI", 10))
        err.pack()

        def do_reset():
            uname, np1, np2 = u.get().strip(), p1.get(), p2.get()
            users = load_users()
            if uname not in users:
                err.configure(text="✗ Username not found.")
                return
            if len(np1) < 6:
                err.configure(text="⚠ Password must be at least 6 characters.")
                return
            if np1 != np2:
                err.configure(text="✗ Passwords do not match.")
                return
            users[uname]["password"] = _hash(np1)
            save_users(users)
            messagebox.showinfo("Done", "Password reset successfully! ✅", parent=win)
            win.destroy()

        ctk.CTkButton(win, text="Reset Password", height=40, fg_color=GOLD_MID, hover_color=GOLD_DARK, text_color=DARK_BG, font=("Segoe UI", 11, "bold"), command=do_reset).pack(padx=30, fill="x", pady=12)

    def _new_account(self):
        win = ctk.CTkToplevel(self)
        win.title("Create Account")
        win.geometry("380x360")
        win.configure(fg_color=DARK_BG)
        win.grab_set()

        ctk.CTkLabel(win, text="Create Account", font=("Georgia", 16, "bold"), text_color=GOLD_LIGHT).pack(pady=(24, 4))
        ctk.CTkLabel(win, text="Set up your extraction credentials", font=("Segoe UI", 10), text_color=GREY_TEXT).pack(pady=(0, 16))

        u = ctk.CTkEntry(win, placeholder_text="Username", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        u.pack(padx=30, fill="x", pady=4)
        em = ctk.CTkEntry(win, placeholder_text="Email (optional)", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        em.pack(padx=30, fill="x", pady=4)
        p1 = ctk.CTkEntry(win, placeholder_text="Password", show="●", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        p1.pack(padx=30, fill="x", pady=4)
        p2 = ctk.CTkEntry(win, placeholder_text="Confirm Password", show="●", height=38, fg_color=CARD_BG, border_color=GOLD_DARK, text_color=WHITE)
        p2.pack(padx=30, fill="x", pady=4)
        err = ctk.CTkLabel(win, text="", text_color=RED_ERR, font=("Segoe UI", 10))
        err.pack()

        def do_create():
            uname, email, np1, np2 = u.get().strip(), em.get().strip(), p1.get(), p2.get()
            users = load_users()
            if not uname:
                err.configure(text="⚠ Username is required."); return
            if uname in users:
                err.configure(text="✗ Username already exists."); return
            if len(np1) < 6:
                err.configure(text="⚠ Password must be at least 6 characters."); return
            if np1 != np2:
                err.configure(text="✗ Passwords do not match."); return
            users[uname] = {"password": _hash(np1), "email": email}
            save_users(users)
            messagebox.showinfo("Success", f"Account '{uname}' created! ✅\nPlease login.", parent=win)
            win.destroy()

        ctk.CTkButton(win, text="Create Account", height=40, fg_color=GOLD_MID, hover_color=GOLD_DARK, text_color=DARK_BG, font=("Segoe UI", 11, "bold"), command=do_create).pack(padx=30, fill="x", pady=12)

# ══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
class Dashboard(ctk.CTkFrame):
    def __init__(self, master, username, on_open_processing):
        super().__init__(master, fg_color=DARK_BG)
        self.username = username
        self.on_open_processing = on_open_processing
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)

        topbar = ctk.CTkFrame(self, fg_color=CARD_BG, corner_radius=0, height=56)
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.grid_columnconfigure(1, weight=1)

        logo_ctk = load_ctk_logo(size=(110, 44))
        if logo_ctk:
            ctk.CTkLabel(topbar, image=logo_ctk, text="").grid(row=0, column=0, padx=16, pady=6)
        else:
            ctk.CTkLabel(topbar, text="RDx Solution", font=("Georgia", 14, "bold"), text_color=GOLD_LIGHT).grid(row=0, column=0, padx=16)

        ctk.CTkLabel(topbar, text=f"Dashboard  ·  v{__version__}", font=("Segoe UI", 13, "bold"), text_color=GOLD_MID).grid(row=0, column=1, sticky="w")

        user_frame = ctk.CTkFrame(topbar, fg_color="transparent")
        user_frame.grid(row=0, column=2, padx=16)
        ctk.CTkLabel(user_frame, text=f"👤 {self.username}", font=("Segoe UI", 10), text_color=GREY_TEXT).pack(side="left", padx=(0, 8))
        ctk.CTkButton(user_frame, text="Logout", width=70, height=28, fg_color=GOLD_DARK, hover_color=RED_ERR, text_color=WHITE, font=("Segoe UI", 9), command=self._logout).pack(side="left")

        ctk.CTkLabel(self, text=f"Welcome back, {self.username} 👋", font=("Georgia", 20, "bold"), text_color=GOLD_LIGHT).grid(row=1, column=0, sticky="w", padx=30, pady=(24, 2))
        ctk.CTkLabel(self, text="Courier Shipping Bill Data Extractor — overview", font=("Segoe UI", 11), text_color=GREY_TEXT).grid(row=2, column=0, sticky="w", padx=30, pady=(0, 16))

        stats = load_stats()
        cards = ctk.CTkFrame(self, fg_color="transparent")
        cards.grid(row=3, column=0, sticky="ew", padx=30)
        for i in range(4): cards.grid_columnconfigure(i, weight=1)

        def stat_card(col, title, value):
            c = ctk.CTkFrame(cards, fg_color=CARD_BG, corner_radius=10, border_color=GOLD_DARK, border_width=1)
            c.grid(row=0, column=col, sticky="nsew", padx=8, pady=4)
            ctk.CTkLabel(c, text=str(value), font=("Segoe UI", 22, "bold"), text_color=GOLD_LIGHT).pack(pady=(16, 0))
            ctk.CTkLabel(c, text=title, font=("Segoe UI", 10), text_color=GREY_TEXT).pack(pady=(2, 16))

        stat_card(0, "Total Files Processed", stats["total_files_processed"])
        stat_card(1, "Total Rows Extracted", stats["total_rows_extracted"])
        stat_card(2, "Total Runs", stats["total_runs"])
        stat_card(3, "Last Run", stats["last_run"] or "—")

        if stats["last_output"]:
            ctk.CTkLabel(self, text=f"📄 Last Excel Output: {stats['last_output']}", font=("Segoe UI", 10), text_color=GREY_TEXT).grid(row=4, column=0, sticky="w", padx=30, pady=(8, 0))

        mon = ctk.CTkFrame(self, fg_color=CARD_BG, corner_radius=10, border_color=GOLD_DARK, border_width=1)
        mon.grid(row=5, column=0, sticky="ew", padx=30, pady=24)
        ctk.CTkLabel(mon, text="System Monitor", font=("Segoe UI", 12, "bold"), text_color=GOLD_MID).pack(pady=(12, 4))
        gframe = ctk.CTkFrame(mon, fg_color="transparent")
        gframe.pack(pady=(0, 14))
        self.cpu_gauge = DialGauge(gframe, label="CPU", size=120)
        self.cpu_gauge.grid(row=0, column=0, padx=20)
        self.ram_gauge = DialGauge(gframe, label="RAM", size=120)
        self.ram_gauge.grid(row=0, column=1, padx=20)
        self._monitoring = True
        self._update_gauges()

        ctk.CTkButton(self, text="▶  Open Processing Window", height=50, corner_radius=8, fg_color=GOLD_MID, hover_color=GOLD_DARK, text_color=DARK_BG, font=("Segoe UI", 14, "bold"), command=self._open).grid(row=6, column=0, sticky="ew", padx=30, pady=(0, 30))

        # ── Update banner (hidden until a newer release is actually found) ──
        self.update_banner = ctk.CTkFrame(self, fg_color=CARD_BG, corner_radius=8, border_color=GOLD_LIGHT, border_width=1)
        self.update_label = ctk.CTkLabel(self.update_banner, text="", font=("Segoe UI", 11), text_color=GOLD_LIGHT)
        self.update_label.pack(side="left", padx=14, pady=10)
        self.update_btn = ctk.CTkButton(self.update_banner, text="Update Now", width=110, height=30, fg_color=GOLD_MID, hover_color=GOLD_DARK, text_color=DARK_BG, command=self._apply_update)
        self.update_btn.pack(side="right", padx=14, pady=10)
        self._update_asset_url = None
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    def _check_update_bg(self):
        has_update, latest_tag, asset_url = updater.check_for_update()
        if has_update:
            self._update_asset_url = asset_url
            self.after(0, lambda: self._show_update_banner(latest_tag))

    def _show_update_banner(self, latest_tag):
        self.update_label.configure(text=f"🔔 New version {latest_tag} available (you have v{__version__})")
        self.update_banner.grid(row=7, column=0, sticky="ew", padx=30, pady=(0, 20))

    def _apply_update(self):
        self.update_btn.configure(state="disabled", text="Updating…")
        threading.Thread(target=self._apply_update_bg, daemon=True).start()

    def _apply_update_bg(self):
        try:
            updater.download_and_apply_update(self._update_asset_url)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror(
                "Update Failed",
                f"Could not apply update automatically: {e}\n\n"
                "You can download the latest version manually from the GitHub Releases page."
            ))
            self.after(0, lambda: self.update_btn.configure(state="normal", text="Update Now"))

    def _update_gauges(self):
        if not getattr(self, "_monitoring", False): return
        try:
            cpu, ram = (psutil.cpu_percent(interval=None), psutil.virtual_memory().percent) if PSUTIL_AVAILABLE else (0, 0)
            self.cpu_gauge.set_value(cpu)
            self.ram_gauge.set_value(ram)
            self.after(1200, self._update_gauges)
        except Exception: pass

    def _open(self):
        self._monitoring = False
        self.on_open_processing(self.username)

    def _logout(self):
        self._monitoring = False
        self.master._show_login()

# ══════════════════════════════════════════════════════════════════════════════
#  CSB EXTRACTOR ENGINE  (real pdfplumber + OCR fallback extraction)
# ══════════════════════════════════════════════════════════════════════════════
DPI_SCALE  = 3          # OCR quality (higher = slower but more accurate)
OCR_CONFIG = "--psm 6"  # Tesseract page segmentation mode

COLUMNS = [
    "File Name",
    "CSB Number",
    "Exchange Rate",
    "Total Taxable Value",
    "Taxable Value Currency",
    "FOB Value (In INR)",
    "Filing Date",
    "EGM Number",
    "EGM Date",
    "HAWB Number",
    "Invoice Number",
    "Invoice Date",
    "Goods Description",
]

# ── pdfplumber — PRIMARY METHOD (text-layer PDFs) ───────────────────────────

def get_all_words(pdf_path):
    """Har word uski (x, y, page) position ke saath extract karta hai."""
    all_words = []
    with pdfplumber.open(pdf_path) as pdf:
        for pg_num, page in enumerate(pdf.pages, 1):
            for w in page.extract_words():
                all_words.append({
                    "text": w["text"],
                    "x0":   w["x0"],
                    "top":  w["top"],
                    "page": pg_num,
                })
    return all_words


def find_value_near_label(words, label_pattern, value_pattern=r".+"):
    """Label ke right ya neeche value dhundta hai."""
    label_re = re.compile(label_pattern, re.IGNORECASE)
    value_re = re.compile(r"^" + value_pattern + r"$", re.IGNORECASE)

    for w in words:
        if label_re.search(w["text"]):
            lpage, ltop = w["page"], w["top"]

            same = sorted(
                [x for x in words
                 if x["page"] == lpage
                 and abs(x["top"] - ltop) <= 5
                 and x["x0"] > w["x0"] + 5
                 and not label_re.search(x["text"])],
                key=lambda x: x["x0"]
            )
            for x in same:
                if value_re.match(x["text"]):
                    return x["text"].strip()

            below = [x for x in words if x["page"] == lpage and 0 < x["top"] - ltop <= 18]
            if below:
                min_top = min(x["top"] for x in below)
                for x in sorted(
                    [x for x in below if abs(x["top"] - min_top) <= 3],
                    key=lambda x: x["x0"]
                ):
                    if value_re.match(x["text"]):
                        return x["text"].strip()
    return ""


def extract_csb_plumber(words):
    """
    CSB Number 2 rows mein hota hai:
      Row 1: CSBNumber:  CSBV_DEL_2026-2027_14
      Row 2:             06_16734
    Joined: CSBV_DEL_2026-2027_14_06_16734
    """
    for w in words:
        if re.search(r"CSBNumber:", w["text"], re.IGNORECASE):
            lpage, ltop = w["page"], w["top"]
            same = sorted(
                [x for x in words if x["page"] == lpage
                 and abs(x["top"] - ltop) <= 5
                 and x["x0"] > w["x0"] + 5],
                key=lambda x: x["x0"]
            )
            part1 = next(
                (x["text"] for x in same if re.match(r"CSBV_", x["text"], re.IGNORECASE)), ""
            )
            if not part1:
                continue
            below = [x for x in words if x["page"] == lpage and 0 < x["top"] - ltop <= 20]
            if below:
                min_top = min(x["top"] for x in below)
                for x in sorted(
                    [x for x in below if abs(x["top"] - min_top) <= 3],
                    key=lambda x: x["x0"]
                ):
                    if re.match(r"^\d{2}[_\-]\d+$", x["text"]):
                        return part1 + "_" + x["text"]
            return part1
    return ""


def extract_invoice_fields_plumber(words):
    """
    Invoice section layout:
      Header row: InvoiceNumber:   InvoiceDate:   InvoiceValue(inINR):
      Data row:   NL-102-26-27     13/06/2026     21523.32
    """
    inv_num_lbl = inv_date_lbl = None
    for w in words:
        if re.match(r"InvoiceNumber:", w["text"], re.IGNORECASE):
            inv_num_lbl = w
        if re.match(r"InvoiceDate:", w["text"], re.IGNORECASE):
            inv_date_lbl = w

    result = {"Invoice Number": "", "Invoice Date": ""}
    if not inv_num_lbl:
        return result

    lpage, ltop = inv_num_lbl["page"], inv_num_lbl["top"]
    below = [x for x in words if x["page"] == lpage and 0 < x["top"] - ltop <= 20]
    if not below:
        return result

    min_top  = min(x["top"] for x in below)
    data_row = sorted(
        [x for x in below if abs(x["top"] - min_top) <= 3],
        key=lambda x: x["x0"]
    )

    for x in data_row:
        if re.match(r"[A-Za-z][A-Za-z0-9\-\/\.]+", x["text"]) and abs(x["x0"] - inv_num_lbl["x0"]) < 80:
            result["Invoice Number"] = x["text"].strip()
            break

    if inv_date_lbl:
        for x in data_row:
            if re.match(r"\d{2}/\d{2}/\d{4}", x["text"]) and abs(x["x0"] - inv_date_lbl["x0"]) < 80:
                result["Invoice Date"] = x["text"].strip()
                break

    return result


def extract_all_items_plumber(words):
    """Multiple item rows: Goods Description + Total Taxable Value."""
    goods_list = []
    taxval_list = []

    goods_re  = re.compile(r"GoodsDescription:", re.IGNORECASE)
    taxval_re = re.compile(r"TotalTaxableValue:", re.IGNORECASE)

    for w in words:
        lpage, ltop = w["page"], w["top"]
        if goods_re.search(w["text"]):
            same = sorted(
                [x for x in words if x["page"] == lpage
                 and abs(x["top"] - ltop) <= 5
                 and x["x0"] > w["x0"] + 5],
                key=lambda x: x["x0"]
            )
            val = " ".join(x["text"] for x in same).strip()
            if val:
                goods_list.append(val)

        if taxval_re.search(w["text"]):
            val = find_value_near_label(words, r"TotalTaxableValue:", r"[\d,\.]+")
            if val and val not in taxval_list:
                taxval_list.append(val)

    n = max(len(goods_list), len(taxval_list), 1)
    return [
        (goods_list[i] if i < len(goods_list) else "",
         taxval_list[i] if i < len(taxval_list) else "")
        for i in range(n)
    ]


def parse_with_plumber(pdf_path):
    """pdfplumber se saare fields extract karta hai."""
    words = get_all_words(pdf_path)

    inv = extract_invoice_fields_plumber(words)
    items = extract_all_items_plumber(words)

    base = {
        "CSB Number":             extract_csb_plumber(words),
        "Filing Date":            find_value_near_label(words, r"FillingDate:",          r"\d{2}/\d{2}/\d{4}"),
        "HAWB Number":            find_value_near_label(words, r"HAWBNumber:",           r"[A-Z0-9]+"),
        "EGM Number":             find_value_near_label(words, r"EGMNumber:",            r"\d+"),
        "EGM Date":               find_value_near_label(words, r"EGMDate:",              r"\d{2}/\d{2}/\d{4}"),
        "Invoice Number":         inv["Invoice Number"],
        "Invoice Date":           inv["Invoice Date"],
        "Taxable Value Currency": find_value_near_label(words, r"TaxableValueCurrency:", r"[A-Z]{3}"),
        "FOB Value (In INR)":     find_value_near_label(words, r"FOBValue\(InINR\):",   r"[\d,\.]+"),
        "Exchange Rate":          find_value_near_label(words, r"ExchangeRate:",         r"[\d,\.]+"),
    }
    return base, items


# ── pytesseract — FALLBACK METHOD (scanned / image-only PDFs) ──────────────

def ocr_pdf_text(pdf_path):
    """PyMuPDF + Tesseract se full text extract karta hai."""
    doc = fitz.open(pdf_path)
    pages = []
    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(DPI_SCALE, DPI_SCALE))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        pages.append(pytesseract.image_to_string(img, config=OCR_CONFIG))
    doc.close()
    return "\n".join(pages)


def extract_csb_ocr(text):
    m = re.search(r"CSB\s*Number[:\s|]+.*?\n\s*([\d\s_]+)", text, re.IGNORECASE)
    if m:
        return re.sub(r"\s+", "_", m.group(1).strip()).strip("_")
    return ""

def extract_filling_date_ocr(text):
    m = re.search(r"Filling\s*Date[:\s|]+([\d]{2}/[\d]{2}/[\d]{4})", text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def extract_hawb_ocr(text):
    m = re.search(r"HAWB\s*Number[:\s|]+([1I][ZG0-9A-Z/\-]+)", text, re.IGNORECASE)
    if m:
        val = re.sub(r"[^A-Z0-9]", "", m.group(1).upper())
        return val[:18]
    return ""

def extract_egm_ocr(text):
    m_no = re.search(r"EGM\s*Number[:\s|]*([\d]+)", text, re.IGNORECASE)
    m_dt = re.search(r"EGM\s*Date[:\s|]*([\d]{2}/[\d]{2}/[\d]{4})", text, re.IGNORECASE)
    return (
        m_no.group(1).strip() if m_no else "",
        m_dt.group(1).strip() if m_dt else "",
    )

def extract_invoice_ocr(text):
    m = re.search(
        r"Invoice\s*Value\s*\(in\s*INR\)[^\n]*\n\s*"
        r"([A-Z]{2,4}-[\d\s/\\\-]+?)\s+([\d]{2}/[\d]{2}/[\d]{4})",
        text, re.IGNORECASE | re.DOTALL
    )
    if m:
        inv_no = re.sub(r"\s*/\s*$", "7", m.group(1).strip())
        inv_no = re.sub(r"\s+", "", inv_no)
        inv_no = re.sub(r"-+", "-", inv_no)
        return inv_no, m.group(2).strip()

    m_no = re.search(r"Invoice\s*Number[:\s|]+([A-Z]{2,4}-[\d\-]+)", text, re.IGNORECASE)
    m_dt = re.search(r"Invoice\s*Date[:\s|]+([\d]{2}/[\d]{2}/[\d]{4})", text, re.IGNORECASE)
    return (
        m_no.group(1).strip() if m_no else "",
        m_dt.group(1).strip() if m_dt else "",
    )

def extract_fob_inr_ocr(text):
    m = re.search(r"FOB\s*Value\s*\(In\s*INR\)[:\s|]*([\d,]+\.?\d*)", text, re.IGNORECASE)
    return m.group(1).replace(",", "").strip() if m else ""

def extract_exchange_rate_ocr(text):
    m = re.search(r"FOB\s*Exchange\s*Rate[^|]*\|\s*([\d.]+)", text)
    if not m:
        m = re.search(r"FOB\s*Exchange\s*Rate[^\d\n]*([\d.]+)", text, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def extract_currency_ocr(text):
    m = re.search(r"FOB\s*Currency[^|]*\|\s*([A-Z]{3})", text)
    return m.group(1).strip() if m else "USD"

def extract_items_ocr(text):
    goods  = [re.sub(r"\s+", " ", g).strip()
              for g in re.findall(r"Goods\s*Description[:\s|]+([^\n]{3,80})", text, re.IGNORECASE)]
    taxval = [v.replace(",", "").strip()
              for v in re.findall(r"Total\s*Taxable\s*Value[:\s|]+([\d.,]+)", text, re.IGNORECASE)]
    n = max(len(goods), len(taxval), 1)
    return [(goods[i] if i < len(goods) else "",
             taxval[i] if i < len(taxval) else "") for i in range(n)]


def parse_with_ocr(pdf_path):
    """OCR fallback – scanned PDFs ke liye."""
    text = ocr_pdf_text(pdf_path)
    egm_no, egm_dt   = extract_egm_ocr(text)
    inv_no, inv_dt   = extract_invoice_ocr(text)
    items            = extract_items_ocr(text)

    base = {
        "CSB Number":             extract_csb_ocr(text),
        "Filing Date":            extract_filling_date_ocr(text),
        "HAWB Number":            extract_hawb_ocr(text),
        "EGM Number":             egm_no,
        "EGM Date":               egm_dt,
        "Invoice Number":         inv_no,
        "Invoice Date":           inv_dt,
        "Taxable Value Currency": extract_currency_ocr(text),
        "FOB Value (In INR)":     extract_fob_inr_ocr(text),
        "Exchange Rate":          extract_exchange_rate_ocr(text),
    }
    return base, items


def has_text_layer(pdf_path, min_chars=50):
    """
    pdfplumber se check karta hai ki PDF mein actual text hai ya nahi.
    Scanned PDFs mein text layer nahi hota.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total = sum(
                len(page.extract_text() or "")
                for page in pdf.pages[:3]
            )
        return total >= min_chars
    except Exception:
        return False


def parse_bill(pdf_path):
    """
    Auto-detect karke correct method choose karta hai:
      1. pdfplumber  — text layer wale PDFs (fast, accurate)
      2. pytesseract — scanned / image PDFs  (slow, needs Tesseract)
    Returns (rows, method).
    """
    fname = os.path.basename(pdf_path)

    if has_text_layer(pdf_path):
        method = "plumber"
        base, items = parse_with_plumber(pdf_path)
    elif OCR_AVAILABLE:
        method = "ocr"
        base, items = parse_with_ocr(pdf_path)
    else:
        raise RuntimeError(
            "PDF mein text layer nahi hai aur OCR libraries install nahi hain.\n"
            "Run: pip install pymupdf pytesseract pillow"
        )

    rows = []
    for goods, taxval in items:
        row = {"File Name": fname, **base,
               "Goods Description":   goods,
               "Total Taxable Value": taxval}
        rows.append(row)

    return rows, method

# ══════════════════════════════════════════════════════════════════════════════
#  PROCESSING WINDOW
# ══════════════════════════════════════════════════════════════════════════════
class ProcessingWindow(ctk.CTkFrame):
    def __init__(self, master, username):
        super().__init__(master, fg_color=DARK_BG)
        self.username = username
        self.selected_files = []
        self._build()

    def _build(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ── Header ────────────────────────────────────────────────────────────
        header = ctk.CTkFrame(self, fg_color=CARD_BG, corner_radius=0, height=56)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(header, text="RDX Extractor Control Hub", font=("Segoe UI", 14, "bold"), text_color=GOLD_LIGHT).grid(row=0, column=0, padx=20, sticky="w")
        ctk.CTkButton(header, text="⬅ Back to Dashboard", width=140, height=30, fg_color=GOLD_DARK, hover_color=GOLD_MID, command=self._back).grid(row=0, column=1, padx=20)

        # ── Control Operations Pane ───────────────────────────────────────────
        ops = ctk.CTkFrame(self, fg_color="transparent")
        ops.grid(row=1, column=0, sticky="ew", padx=20, pady=15)
        
        self.add_btn = ctk.CTkButton(ops, text="📂 Add Shipping Bills (PDF)", font=("Segoe UI", 12, "bold"), fg_color=GOLD_MID, text_color=DARK_BG, hover_color=GOLD_LIGHT, command=self._add_files)
        self.add_btn.pack(side="left", padx=5)
        
        self.clear_btn = ctk.CTkButton(ops, text="🗑 Clear List", font=("Segoe UI", 12), fg_color="#3A2E20", text_color=WHITE, hover_color=RED_ERR, command=self._clear_list)
        self.clear_btn.pack(side="left", padx=5)

        self.run_btn = ctk.CTkButton(ops, text="⚡ Run Extraction Matrix", font=("Segoe UI", 12, "bold"), fg_color=GREEN_OK, text_color=WHITE, hover_color="#66C266", command=self._run_extraction)
        self.run_btn.pack(side="right", padx=5)

        # ── Dynamic Progress Indicator ────────────────────────────────────────
        self.prog_bar = ctk.CTkProgressBar(self, fg_color=CARD_BG, progress_color=GOLD_MID)
        self.prog_bar.grid(row=3, column=0, sticky="ew", padx=25, pady=(5, 5))
        self.prog_bar.set(0)

        # ── Informative Output Area ───────────────────────────────────────────
        self.console_frame = ctk.CTkFrame(self, fg_color=CARD_BG, border_color=GOLD_DARK, border_width=1)
        self.console_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=5)
        
        self.txt = tk.Text(self.console_frame, bg=CARD_BG, fg=CREAM, insertbackground=GOLD_LIGHT, font=("Consolas", 11), bd=0, highlightthickness=0)
        self.txt.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        scr = ctk.CTkScrollbar(self.console_frame, command=self.txt.yview)
        scr.pack(side="right", fill="y")
        self.txt.configure(yscrollcommand=scr.set)
        
        self._log("💡 RDX Solution Engine Standby Mode. Upload Courier Bills to start processing pipeline.\n")

    def _log(self, msg):
        self.txt.insert(tk.END, msg)
        self.txt.see(tk.END)

    def _add_files(self):
        fps = filedialog.askopenfilenames(filetypes=[("Courier Manifest PDF", "*.pdf")])
        if fps:
            for f in fps:
                if f not in self.selected_files:
                    self.selected_files.append(f)
                    self._log(f"📌 Loaded Manifest: {os.path.basename(f)}\n")

    def _clear_list(self):
        self.selected_files.clear()
        self.txt.delete("1.0", tk.END)
        self.prog_bar.set(0)
        self._log("💡 System Cache Cleared. Queue Ready.\n")

    def _back(self):
        self.master._show_dashboard(self.username)

    def _run_extraction(self):
        if not self.selected_files:
            messagebox.showwarning("Empty Context", "Please load source shipping records (PDF format) first.")
            return
        self.add_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        self.clear_btn.configure(state="disabled")
        
        threading.Thread(target=self._proc_thread, daemon=True).start()

    def _proc_thread(self):
        all_rows = []
        tot = len(self.selected_files)
        
        for idx, fp in enumerate(self.selected_files, 1):
            fn = os.path.basename(fp)
            self._log(f"⚡ Analyzing data architecture inside -> {fn}...\n")
            try:
                res, method = parse_bill(fp)
                all_rows.extend(res)
                self._log(f"   ↳ Extracted {len(res)} row(s) via [{method}] method.\n")
            except Exception as e:
                self._log(f"❌ Structural Failure parsing: {fn}. Details: {str(e)}\n")
            self.prog_bar.set(idx / tot)
            time.sleep(0.2)

        if not all_rows:
            self._log("⚠️ Process execution aborted: No structured shipping fields recovered.\n")
            self._reset_buttons()
            return

        # Build highly customized & styled Excel ledger matrix
        self._log("📊 Exporting normalized datasets into professional corporate ledger template...\n")
        out_fp = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workspace", "*.xlsx")])
        if not out_fp:
            out_fp = os.path.join(os.path.expanduser("~"), "Desktop", f"RDX_CSBV_Output_{int(time.time())}.xlsx")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "CSB-V Extracted Data"
            ws.views.sheetView[0].showGridLines = True

            # Elite corporate design palette styling rules
            hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            reg_font = Font(name="Segoe UI", size=10, color="000000")
            hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Classic corporate dark navy
            alt_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Subtle cool grey alternate striping
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )

            # Write high-contrast column structures
            ws.append(COLUMNS)
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Render data records
            for r_idx, data_dict in enumerate(all_rows, start=2):
                row_data = [data_dict.get(c, "") for c in COLUMNS]
                ws.append(row_data)
                
                is_alt = (r_idx % 2 == 0)
                for col_idx in range(1, len(COLUMNS) + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    cell.font = reg_font
                    cell.border = thin_border
                    
                    if is_alt:
                        cell.fill = alt_fill
                        
                    # Custom programmatic alignment masks based on column data content
                    cname = COLUMNS[col_idx - 1]
                    if cname in ["Total Taxable Value", "FOB Value (In INR)", "Exchange Rate"]:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.00'
                    elif "Date" in cname or cname in ["CSB Number", "HAWB Number", "Invoice Number"]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[1].height = 28
            for r in range(2, len(all_rows) + 2):
                ws.row_dimensions[r].height = 20

            # Autofit layout calculation engine boundaries
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(out_fp)
            record_run(tot, len(all_rows), out_fp)
            self._log(f"💥 PIPELINE COMPLETE! Ledger compiled and securely written to:\n 👉 {out_fp}\n")
            messagebox.showinfo("Pipeline Complete", f"Data matrices successfully compiled!\nProcessed: {tot} files\nExtracted: {len(all_rows)} records.")
        except Exception as e:
            self._log(f"❌ File I/O Error saving generated template layout: {str(e)}\n")
            messagebox.showerror("Export Crash", f"Could not compile data array template sheet: {str(e)}")

        self._reset_buttons()

    def _reset_buttons(self):
        self.add_btn.configure(state="normal")
        self.run_btn.configure(state="normal")
        self.clear_btn.configure(state="normal")

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN WINDOW APPLICATION ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("RDX Solution — Courier Shipping Bill Data Extractor v3.4.2")
        self.geometry("900x620")
        self.configure(fg_color=DARK_BG)
        self.resizable(False, False)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.current_frame = None
        
        # Deploy initial Flash splash frame sequence
        self.withdraw()
        splash = SplashScreen(self)
        self.after(2800, lambda: self._terminate_splash(splash))

    def _terminate_splash(self, splash):
        splash._failsafe_close()
        self.deiconify()
        self._show_login()

    def _switch_frame(self, frame_class, *args, **kwargs):
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = frame_class(self, *args, **kwargs)
        self.current_frame.grid(row=0, column=0, sticky="nsew")

    def _show_login(self):
        self._switch_frame(LoginPage, on_login_success=self._show_dashboard)

    def _show_dashboard(self, username):
        self._switch_frame(Dashboard, username=username, on_open_processing=self._show_processing)

    def _show_processing(self, username):
        self._switch_frame(ProcessingWindow, username=username)

if __name__ == "__main__":
    app = App()
    app.mainloop()
